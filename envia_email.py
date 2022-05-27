from time import sleep
import sys
import json
import os
import openpyxl as op
from datetime import date
import locale
from send_email import send_mail
from cliente import Cliente
locale.setlocale(locale.LC_TIME, '')

# Method to read config file settings
with open('./config.json', encoding="utf-8") as config_file:
    config = json.load(config_file)

planilha_base = config['planilha_base']
caminho_raiz_anexos = config['caminho_raiz_anexos']
paths_finais = config['paths_finais'].split(';')
email_de_envio = config['email_de_envio']

body_email = '''
                <br/>
                <p><strong>Att.</strong></p>
                <p><strong>Gabriela Andreoli Pedro</strong><strong>
                <br/strong>Departamento Financeiro
                <br/><img src="https://raw.githubusercontent.com/leonan0/files/main/logoacn.jpg" alt="" />
                <br/>Avenida Rio de Janeiro, 1765 &ndash; Londrina/PR
                <br/>Telefone: (43) 3324-4373
                <br/>E-mail:&nbsp;<a href="mailto:office01@acncon.com.br">financeiro@acncon.com.br</a></p>
'''

MONTH = date.today().strftime("%B").capitalize()
YEAR = str(date.today().year)
PATH = f'{caminho_raiz_anexos}\\{YEAR}\\{MONTH}'
PATH_ANEXOS = paths_finais

path = planilha_base
wb_obj = op.load_workbook(path)

clientes = wb_obj['Clientes']


if MONTH in wb_obj.sheetnames:
    wb_obj.remove(wb_obj[MONTH])
    dados = wb_obj.create_sheet(MONTH)
    dados.append(['emails', 'assunto', 'texto', 'paths', 'qtd_anexos',
                'empresas', 'qtd_empresas', 'data_envio'])


all_paths = []
for i in PATH_ANEXOS:
    path_x = os.path.join(PATH, i)
    files = os.listdir(path_x)
    files_paths = [f"{path_x}\{x}" for x in files]
    all_paths.extend(files_paths)

for k, v in enumerate(all_paths):
    all_paths[k] = v.replace('\\', '/')


registros = []
for r in range(2, clientes.max_row + 1):
    cliente = []
    for i in range(1, clientes.max_column + 1):
        cell = clientes.cell(r, i)
        cliente.append(cell.value)

    if not cliente[1]:
        cliente[1] = ""
    cliente[0] = cliente[0].split(' ')
    cliente[2] = cliente[2].split(';')
    cliente[2] = [x.strip() for x in cliente[2]]

    cli_obj = {
        "email": cliente[0], "nome_cliente": cliente[1], "empresa": cliente[2], "all_paths": all_paths, "assinatura": body_email
    }
    cli = Cliente(**cli_obj)
    registros.append(cli)

for r in registros:
    dados.append((r.to_tuple()))

wb_obj.save(path)
wb_obj.close()

if sys.argv[1] == '1':
    print('Envio de email desligado.')
    print('FIM')
else:
    for r in registros:
        send_mail(email_de_envio, **r.to_dict())
print(f'Geradas {len(registros)} linhas')
sleep(15)